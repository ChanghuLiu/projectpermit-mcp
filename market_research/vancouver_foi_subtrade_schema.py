"""Inspect only schema metadata from Vancouver public FOI permit workbook.

No data-row values are printed. This probe exists to discover sheet names and
column headers safely before designing aggregate sub-trade workload analysis.
"""
from __future__ import annotations

import argparse
import io
import json
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

SOURCE_URL = "https://vancouver.ca/files/cov/2024-671-release2.XLSX"


def _read_payload(input_file: Path | None) -> bytes:
    if input_file is not None:
        return input_file.read_bytes()

    request = urllib.request.Request(
        SOURCE_URL,
        headers={
            "User-Agent": "Mozilla/5.0 ProjectPermit public-market-research/1.0",
            "Referer": "https://vancouver.ca/your-government/freedom-of-information.aspx",
        },
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", type=Path)
    parser.add_argument("--output", type=Path, default=Path("vancouver_foi_schema.json"))
    args = parser.parse_args()

    payload = _read_payload(args.input_file)
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    result = {"source_url": SOURCE_URL, "sheets": []}

    for sheet in workbook.worksheets:
        # Find the first non-empty row and treat it as the candidate header row.
        header_row_index = None
        headers = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                header_row_index = idx
                headers = values
                break
        result["sheets"].append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "candidate_header_row": header_row_index,
                "headers": headers,
            }
        )

    rendered = json.dumps(result, indent=2, ensure_ascii=False)
    args.output.write_text(rendered + "\n", encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()
